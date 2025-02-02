import json
import logging
import os

from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from app.config.config import settings

logger = logging.getLogger(__name__)

# Функция загрузки конфигурации
def load_config(file_path):
    with open(file_path, "r", encoding="utf-8") as f:
        return json.load(f)

# Создаем пустую структуру для хранения данных
def create_data_structure(config):
    data_structure = {}
    for section in config:
        for sub_header in section.get("sub_headers", []):
            data_structure[sub_header["data_variable"]] = None
    return data_structure

data = create_data_structure(load_config(settings.PATH_TO_CONFIG_RESULT_TABLE))

def create_excel_from_config(config, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    filename = f"{os.path.dirname(filename)}/{os.path.splitext(os.path.basename(filename))[0]}-output.xlsx"

    # Заполняем заголовки
    current_row = 1
    current_col = 1
    try:
        for section in config:
            # Верхний уровень заголовков
            if section["title"]:
                col_span = section["col_span"]
                ws.merge_cells(
                    start_row=current_row,
                    start_column=current_col,
                    end_row=current_row,
                    end_column=current_col + col_span - 1,
                )
                ws.cell(row=current_row, column=current_col).value = section["title"]

            # Подзаголовки
            for sub_header in section["sub_headers"]:
                col_span = sub_header["col_span"]
                row_span = sub_header["row_span"]
                if section["title"] == "":
                    ws.cell(row=current_row, column=current_col).value = sub_header["title"]
                    ws.merge_cells(
                        start_row=current_row,
                        start_column=current_col,
                        end_row=current_row + row_span,
                        end_column=current_col + col_span - 1,
                    )
                else:
                    ws.cell(row=current_row + 1, column=current_col).value = sub_header["title"]
                    ws.merge_cells(
                    start_row=current_row + 1,
                    start_column=current_col,
                    end_row=current_row + row_span,
                    end_column=current_col + col_span - 1,
                )
                current_col += col_span

                # Сохраняем файл
                wb.save(filename)

        logger.info(f"Excel file created successfully: filename")
        return filename
    except Exception as e:
        logger.error(f"Error creating Excel file: {str(e)}")

def append_df_to_excel(filename, df, sheet_name="Sheet1", start_row=None, start_col=None):
    """
    Добавление DataFrame в Excel файл начиная с определенной строки и столбца.
    Если файл не существует, он будет создан.
    Если лист не существует, он будет создан.

    :param filename: Путь до Excel файла
    :param df: DataFrame для записи
    :param sheet_name: Имя листа для записи данных
    :param start_row: Начальная строка для записи данных DataFrame
    :param start_col: Начальный столбец для записи данных DataFrame
    """

    # Загрузить книгу и выбрать лист
    book = load_workbook(filename)
    if sheet_name not in book.sheetnames:
        book.create_sheet(sheet_name)
    sheet = book[sheet_name]

    # Пересчитать start_row в соответствии с содержимым листа, если не указано
    if start_row is None:
        start_row = sheet.max_row + 1

    # Преобразовать DataFrame в строки Excel и записать данные
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
        for c_idx, value in enumerate(row, start=start_col if start_col is not None else 1):
            sheet.cell(row=r_idx, column=c_idx, value=value)

    book.save(filename)
    book.close()



