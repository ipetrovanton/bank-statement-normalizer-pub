import logging
import re
from collections import Counter
from enum import Enum, auto
from io import BytesIO
from typing import List, Dict, Any, Optional, Tuple

import numpy as np
import openpyxl
import pandas as pd

from app.config.result_messages import ResultMessages
from app.constants.constants import SEQ_NUMBER, DOCUMENT_OPERATION_DATE, DOCUMENT_TYPE_CODE, DOCUMENT_NUMBER, \
    DOCUMENT_DATE, \
    CORESPONDENT_ACCOUNT_NUMBER, PAYER_OR_RECIPIENT_BANK, PAYER_OR_RECIPIENT_NAME, ACCOUNT_NUMBER, DEBIT_AMOUNT, \
    CREDIT_AMOUNT, PAYMENT_PURPOSE, PAYER_OR_RECIPIENT_INN, PAYER_OR_RECIPIENT_KPP, BANK_BIK, DEBTOR_ACCOUNT_NUMBER, \
    DEBTOR_BANK_NAME, DEBTOR_NAME, CURRENCY_CODE
from app.constants.regex_patterns import PAYMENT_PURPOSE_PATTERN, PAYER_OR_RECIPIENT_KPP_PATTERN, \
    PAYER_OR_RECIPIENT_INN_PATTERN, BANK_BIK_PATTERN, OPERATION_DATE_PATTERN, DOCUMENT_TYPE_CODE_PATTERN, \
    DOCUMENT_NUMBER_PATTERN, DOCUMENT_DATE_PATTERN, CORESPONDENT_ACCOUNT_NUMBER_PATTERN, \
    PAYER_OR_RECIPIENT_BANK_PATTERN, SEQ_NUMBER_PATTERN, DEBIT_AMOUNT_PATTERN, \
    CREDIT_AMOUNT_PATTERN, PAYER_OR_RECIPIENT_NAME_PATTERN, ACCOUNT_NUMBER_PATTERN, DEBTOR_ACCOUNT_NUMBER_PATTERN, \
    DEBTOR_BANK_NAME_PATTERN, DEBTOR_BANK_NAME_VALUE_PATTERN, DEBTOR_NAME_PATTERN, DEBTOR_NAME_VALUE_PATTERN, \
    CURRENCY_CODE_PATTERN, DEBTOR_ACCOUNT_NUMBER_VALUE_PATTERN, CURRENCY_CODE_VALUE_PATTERN

logger = logging.getLogger(__name__)  # Логгер на уровне модуля

class ErrorSeverity(Enum):
    """Enum to define error severity levels"""
    CRITICAL = auto()  # Stops further processing
    WARNING = auto()   # Allows processing to continue
    INFO = auto()      # Informational message

class ProcessingError:
    """Enhanced error class for preprocessing errors"""
    def __init__(
        self,
        code: int,
        message: str,
        severity: ErrorSeverity = ErrorSeverity.WARNING,
        details: Optional[Dict[str, Any]] = None
    ):
        self.code = code
        self.message = message
        self.severity = severity
        self.details = details or {}

    def to_dict(self) -> Dict[str, Any]:
        """Convert error to dictionary representation"""
        return {
            "code": self.code,
            "message": self.message,
            "severity": self.severity.name,
            "details": self.details
        }

def count_of_filled_sheets_in_wb(wb) -> Tuple[int, List[ProcessingError]]:
    """
    Подсчет заполненных листов в рабочей книге с обработкой ошибок

    Returns:
    - int: Количество заполненных листов
    - List[ProcessingError]: Список ошибок обработки
    """
    errors = []
    filled_sheets_count = 0
    try:
        # Перебор всех листов в книге
        for sheet in wb.worksheets:
            # Проверка, есть ли в листе данные
            if sheet.max_row > 1 or (sheet.max_row == 1 and sheet.max_column > 1):
                # Если есть хотя бы одна строка и одна колонка (кроме случая одной ячейки A1),
                # считаем лист заполненным.
                filled_sheets_count += 1
    except Exception as e:
        sheet_count_error = ProcessingError(
            code=500,
            message=f"Ошибка подсчета заполненных листов: {str(e)}",
            severity=ErrorSeverity.CRITICAL,
            details={"exception": str(e)}
        )
        errors.append(sheet_count_error)
    return filled_sheets_count, errors

def parse_xlsx_to_df(file_path) -> Tuple[pd.DataFrame, List[ProcessingError]]:
    """
    Enhanced method to parse XLSX file with comprehensive error handling

    Returns:
    - DataFrame: Processed dataframe
    - List of ProcessingErrors: Warnings and errors encountered during processing
    """
    errors = []
    df = pd.DataFrame()

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
    except Exception as e:
        logger.error(f"Error processing file: {str(e)}")
        critical_error = ProcessingError(
            code=500,
            message=f"Failed to read file: {file_path}",
            severity=ErrorSeverity.CRITICAL,
            details={"exception": str(e)}
        )
        errors.append(critical_error)
        return df, errors

    filled_sheets_in_wb, filled_sheets_errors = count_of_filled_sheets_in_wb(wb)
    errors.extend(filled_sheets_errors)

    if filled_sheets_in_wb == 0:
        logger.info(f"File: {file_path} is empty!")
        empty_file_error = ProcessingError(
            code=400,
            message=f"File: {file_path} is empty!",
            severity=ErrorSeverity.WARNING
        )
        errors.append(empty_file_error)
        return df, errors

    elif filled_sheets_in_wb > 1:
        logger.info(f"Number of filled sheets in file {file_path} is more than 1!")
        multiple_sheets_error = ProcessingError(
            code=400,
            message=f"Number of filled sheets in file {file_path} is more than 1!",
            severity=ErrorSeverity.WARNING
        )
        errors.append(multiple_sheets_error)
        return df, errors

    # Process single sheet
    worksheet_name = wb.sheetnames[0]
    df = pd.read_excel(file_path, engine='openpyxl', sheet_name=worksheet_name, header=None)

    # Process the dataframe
    df, processing_errors = processing_file_df(df)

    # Combine and return errors
    errors.extend(processing_errors)

    return df, errors

def processing_file_df(df) -> Tuple[pd.DataFrame, List[ProcessingError]]:
    """
    Enhanced method to process DataFrame with comprehensive error handling

    Returns:
    - DataFrame: Processed dataframe
    - List of ProcessingErrors: Warnings and errors encountered during processing
    """
    errors = []

    try:
        # Парсинг DataFrame на области поиска, соответствующие логическим блокам размещения информации
        dfs = parse_df_to_section(df)
    except Exception as e:
        parsing_error = ProcessingError(
            code=500,
            message=f"Ошибка парсинга DataFrame на логические блоки: {str(e)}",
            severity=ErrorSeverity.CRITICAL,
            details={"exception": str(e)}
        )
        errors.append(parsing_error)
        return pd.DataFrame(), errors

    # Проверка количества возвращенных областей поиска
    if len(dfs) == 2:
        bank_statement_df = dfs[1][0]
        headers_count = dfs[1][1]

        bank_statement_df, bank_statement_errors = processing_bank_statement_section(bank_statement_df, headers_count)
        if bank_statement_errors is not None and [error for error in bank_statement_errors if error.severity == ErrorSeverity.CRITICAL]:
            return bank_statement_df, bank_statement_errors
        if bank_statement_errors is not None:
            errors.extend(bank_statement_errors)

        # Значения для поиска заголовков о соответствующих им значений
        general_info_headers = [
            DEBTOR_ACCOUNT_NUMBER,
            CURRENCY_CODE,
            DEBTOR_BANK_NAME,
            DEBTOR_NAME
        ]

        bank_statement_column_names_set = set(bank_statement_df.columns)
        common_info_column_names_set = set(general_info_headers)

        # Находим общие заголовки
        common_col = bank_statement_column_names_set.intersection(common_info_column_names_set)

        # Если общие заголовки полностью совпадают с заголовками общей банковской информации
        if common_col == common_info_column_names_set:
            return bank_statement_df, errors

        # Если общие заголовки не полностью совпадают
        else:
            common_info_column_names_set_diff = common_info_column_names_set.difference(bank_statement_column_names_set)

            # Обработка блока общей информации
            general_bank_info_df = dfs[0]
            result = processing_general_bank_info_section(general_bank_info_df, common_info_column_names_set_diff)
            errors.extend(result[1])
            common_col_value = result[0]

            # Объединение логических блоков в единый DataFrame
            for col, val in common_col_value.items():
                bank_statement_df[col] = val

            return bank_statement_df, errors

    # Обработка случая с одним блоком
    elif len(dfs) == 1:
        # Обработка блока общей информации
        general_bank_info_df = dfs[0]
        common_col_value, general_info_errors = processing_general_bank_info_section(general_bank_info_df, None)
        errors.extend(general_info_errors)

        single_block_warning = ProcessingError(
            code=300,
            message="Не удалось найти блок информации о совершенных операциях в файле!",
            severity=ErrorSeverity.WARNING
        )
        errors.append(single_block_warning)

        bank_statement_df = pd.DataFrame.from_dict(common_col_value, orient='index').T
        return bank_statement_df, errors

    # Вернулся пустой список
    else:
        no_sections_error = ProcessingError(
            code=400,
            message="Не удалось выделить логические области в файле!",
            severity=ErrorSeverity.CRITICAL
        )
        errors.append(no_sections_error)
        return pd.DataFrame(), errors

def processing_bank_statement_section(df_of_bank_statement, headers_count) -> Tuple[pd.DataFrame, List[ProcessingError]]:
    """
    Обработка блока информации о совершенных операциях

    Returns:
    - DataFrame: Обработанный DataFrame
    - List[ProcessingError]: Список ошибок обработки
    """
    errors = []

    # Чистка DataFrame
    df_of_bank_statement, cleaning_errors = clean_dataframe(df_of_bank_statement)
    errors.extend(cleaning_errors)

    # Корректировка заголовков DataFrame, замена на типовые заголовки
    try:
        df_of_bank_statement, origin_headers, header_correction_errors = correct_df_headers(df_of_bank_statement, headers_count)
        if header_correction_errors is not None and [error for error in header_correction_errors if error.severity == ErrorSeverity.CRITICAL]:
            return df_of_bank_statement, header_correction_errors
    except Exception as e:
        header_correction_error = ProcessingError(
            code=400,
            message=f"Ошибка корректировки заголовков: {str(e)}",
            severity=ErrorSeverity.WARNING,
            details={"exception": str(e)}
        )
        errors.append(header_correction_error)
        return df_of_bank_statement, errors

    # Удаление строк с агрегирующими значениями
    df_of_bank_statement, agregate_rows_errors = clean_agregate_rows(df_of_bank_statement)
    errors.extend(agregate_rows_errors)

    # Удаление заголовков для выписок, представляющих сведения одновременно по нескольким счетам
    df_of_bank_statement, headers_in_bank_statement_errors = clean_headers_in_bank_statement(df_of_bank_statement, origin_headers)
    errors.extend(headers_in_bank_statement_errors)

    # Удаление строк с порядковыми номерами столбцов
    df_of_bank_statement = clean_sequential_columns_numbers(df_of_bank_statement)

    if len(df_of_bank_statement) == 0:
        no_operations_warning = ProcessingError(
            code=300,
            message="В выписке нет операций!",
            severity=ErrorSeverity.WARNING
        )
        errors.append(no_operations_warning)

    return df_of_bank_statement, errors

def processing_general_bank_info_section(general_bank_info_df, common_info_column_names_set_diff=None) -> (Dict[str, Any], List[ProcessingError]):
    """
    Обработка блока общей информации

    Returns:
    - Dict: Словарь общей банковской информации
    - List[ProcessingError]: Список ошибок обработки
    """
    errors = []

    # Значения для поиска заголовков о соответствующих им значений
    general_info_correlation = {
        DEBTOR_ACCOUNT_NUMBER_PATTERN: [DEBTOR_ACCOUNT_NUMBER, DEBTOR_ACCOUNT_NUMBER_VALUE_PATTERN],
        CURRENCY_CODE_PATTERN: [CURRENCY_CODE, CURRENCY_CODE_VALUE_PATTERN],
        DEBTOR_BANK_NAME_PATTERN: [DEBTOR_BANK_NAME, DEBTOR_BANK_NAME_VALUE_PATTERN],
        DEBTOR_NAME_PATTERN: [DEBTOR_NAME, DEBTOR_NAME_VALUE_PATTERN]
    }

    if common_info_column_names_set_diff is not None:
        # Оставляем в словаре только те пары, ключи которых есть в множестве
        general_info_correlation = {
            key: value
            for key, value in general_info_correlation.items()
            if value[0] in common_info_column_names_set_diff
        }

    try:
        # Чистка DataFrame
        general_bank_info_df, cleaning_errors = clean_dataframe(general_bank_info_df)
        errors.extend(cleaning_errors)

        # Возврат словаря 'Название заголовка': 'Значение заголовка'
        common_col_value, general_info_errors = search_general_bank_info(general_bank_info_df, general_info_correlation)
        errors.extend(general_info_errors)
    except Exception as e:
        general_info_error = ProcessingError(
            code=400,
            message=f"Ошибка обработки общей банковской информации: {str(e)}",
            severity=ErrorSeverity.WARNING,
            details={"exception": str(e)}
        )
        errors.append(general_info_error)
        common_col_value = {}

    return common_col_value, errors

def compute_fullness_of_row(df):
    # Определяем пустой словарь
    fullness_of_rows = {}
    for row in df.index:
        fill_or_empty_cells_in_row = []
        for col in range(len(df.columns)):
            cell_value = df.iat[row, col]
            if pd.isna(cell_value):
                # Если значение ячейки пустое, то переходим к следующей ячейке
                fill_or_empty_cells_in_row.append(False)
                continue
            else:
                fill_or_empty_cells_in_row.append(True)
        fullness_of_rows[row] = sum(fill_or_empty_cells_in_row)
    return fullness_of_rows

# get_key_of_most_frequent_value_in_dict функция определения минимального/максимального ключа словаря
# среди пар, у которых значение равно самому часто встречающемуся значению в словаре
# возвращает ключ словаря
def get_key_of_most_frequent_value_in_dict(some_dict, value_limit, min_flag = True) -> (int|None):
    # some_dict {индекс строки: количество заполненных ячеек}
    # Подсчет количества каждого значения в словаре, но только если значение между 10 и 18
    # количество заполненных ячеек: количество повторений
    value_counts = Counter(value for value in some_dict.values() if value_limit[0] < value < value_limit[1])
    if not value_counts:
        return None
    # Нахождение самого часто встречающегося значения
    # most_common_value = value_counts.most_common(1)[0][0]
    most_common_value = value_counts.most_common()
    second_elements = {x[1] for x in most_common_value}
    if len(second_elements) == 1:
        if min_flag:
            most_common_value = most_common_value[0][0]
        else:
            most_common_value = most_common_value[-1][0]
    else:
        most_common_value = most_common_value[0][0]
    # Выбор всех ключей, у которых значение равно самому часто встречающемуся
    keys_with_most_common_value = [key for key, value in some_dict.items() if value == most_common_value]
    # Нахождение минимального ключа
    if min_flag:
        key_with_most_common_value = min(keys_with_most_common_value)
    else:
        key_with_most_common_value = max(keys_with_most_common_value)
    return key_with_most_common_value


# parse_df_to_section функция парсинга DataFrame на 2 логические области:
# 1 - Общая информация Выписки
# 2 - Информация об операциях, размер заголовка
def parse_df_to_section(df):
    found_dfs = []
    # headers_rows_count = 0
    df, cleaning_errors = clean_dataframe(df)
    errors = []
    # Высчитываем количество заполненных ячеек в каждой строке
    fullness_of_rows = compute_fullness_of_row(df)
    # Устанавливаем примерный диапазон количества колонок в таблице
    approximate_quantity_of_table_col = (9, 19)
    # Получаем индекс строки примерного расположения начала размещения данных таблицы (блока информации о совершенных операции)
    min_row_index = get_key_of_most_frequent_value_in_dict(fullness_of_rows, approximate_quantity_of_table_col)
    # Если индекс строки примерного расположения начала размещения данных таблицы, не установлен
    if min_row_index is None:
        logger.info("В файле не удалось найти таблицу, содержащую информацию о совершенных операциях!")
        # Возвращаем список с одним DataFrame, для поиска общей банковской информации
        found_dfs.append(df)
        return found_dfs
    # Определяем наиболее популярные заголовки, которые встречаются только в блоке информации о совершенных операция,
    # задаем список паттернов для этих заголовков
    most_popular_headers = [
        OPERATION_DATE_PATTERN,
        DOCUMENT_TYPE_CODE_PATTERN,
        CORESPONDENT_ACCOUNT_NUMBER_PATTERN,
        DEBIT_AMOUNT_PATTERN,
        CREDIT_AMOUNT_PATTERN,
        PAYMENT_PURPOSE_PATTERN
    ]
    # Формируем общее регулярное выражение
    regex = r'|'.join(most_popular_headers)
    # Определяем область поиска заголовков таблицы
    search_df = df.iloc[:min_row_index+1]
    # Осуществляем поиск популярных заголовков, возвращаем координаты
    headers_found_coordinates = find_regex_in_df(search_df, regex)
    # Если заголовки не найдены
    if not headers_found_coordinates:
        logger.info("Заголовки таблицы, содержащей информацию о совершенных операциях, не идентифицированы!")
        # Возвращаем список с одним DataFrame, для поиска общей банковской информации
        found_dfs.append(df)
        return found_dfs
    else:
        # Извлечение первого элемента из каждого кортежа и подсчет их количества
        # т.е. подсчет встречаемости заголовков в каждой строке
        rows_of_found_headers = Counter(t[0] for t in headers_found_coordinates)
        # Если заголовки, принадлежащие блоку информации о совершенных операциях найдены в одной строке
        if len(rows_of_found_headers) == 1:
            headers_rows_count = 1
            # Получение ключа из словаря
            row_parting  = next(iter(rows_of_found_headers))
        # Если заголовки, принадлежащие блоку информации о совершенных операциях найдены в нескольких строках
        else:
            headers_rows_count = 2
            step = 1
            # Находим индекс строки, в которой наибольшее количество совпадений по заголовкам
            max_row_index_of_headers_presence = max(rows_of_found_headers, key=rows_of_found_headers.get)
            # Проверяем присутствие индексов строк выше и ниже
            rows_range_of_headers_presence = get_unit_from_key(rows_of_found_headers, max_row_index_of_headers_presence, step)
            # Если вернулось 2 индекса строк
            if len(rows_range_of_headers_presence) == 2:
                # Получение значений по выбранным ключам
                rows_range = {key: rows_of_found_headers[key] for key in rows_range_of_headers_presence}
                # Удаляем индекс строки, в которой наименьшее количество совпадений по заголовкам
                min_value_key = min(rows_range, key=rows_range.get)
                del rows_range[min_value_key]
            # Добавляем индекс строки, в которой наибольшее количество совпадений по заголовкам
            rows_range_of_headers_presence.append(max_row_index_of_headers_presence)
            # Получение значений по выбранным ключам
            selected_values = {key: rows_of_found_headers[key] for key in rows_range_of_headers_presence}
            # Нахождение минимального индекса строки из отобранных строк,
            # предположительно содержащих заголовки блока информации о совершенных операциях
            row_parting = min(selected_values.keys())
        # Получаем индекс строки примерного расположения конца размещения данных таблицы (блока информации о совершенных операция)
        max_row_index = get_key_of_most_frequent_value_in_dict(fullness_of_rows, approximate_quantity_of_table_col, False)
        # Cписок логических блоков DataFrame, и размерности заголовка для таблицы с информацией о совершенных операциях
        found_dfs = [df.iloc[:row_parting], [df.iloc[row_parting:max_row_index+1], headers_rows_count]]
    return found_dfs

# get_unit_from_key - функция получения ключей с заданным шагом от передаваемого ключа
# возвращает список найденных ключей
def get_unit_from_key(some_dict, start_key, unit):
    # Проверка наличия ключ start_key
    if start_key not in some_dict:
        raise ValueError('Переданного ключа нет в словаре')
    selected_keys = []
    # Проверка наличия ключей max_key + unit и max_key - unit
    key_plus_one_exists = start_key + unit in some_dict
    key_minus_one_exists = start_key - unit in some_dict
    # Если присутствуют оба
    if key_plus_one_exists and key_minus_one_exists:
        selected_keys.append(start_key + unit)
        selected_keys.append(start_key - unit)
    # Если присутствует положительный
    elif key_plus_one_exists and not key_minus_one_exists:
        selected_keys.append(start_key + unit)
    # Если присутствует отрицательный
    elif key_minus_one_exists and not key_plus_one_exists:
        selected_keys.append(start_key - unit)
    return selected_keys

# delete_symbols_from_string функция удаления переданных символов из строки
def delete_symbols_from_string(str_to_clean, chars_to_remove):
    # Создание таблицы перевода
    trans_table = str.maketrans('/', ' ', chars_to_remove)
    # Удаление символов из строки
    cleaned_string = str_to_clean.translate(trans_table)
    return cleaned_string

# find_regex_in_df функция поиска в Dataframe значения, соответствующего переданному регулярному выражению
# возвращает список кортежей с координатами соответствующих регулярному выражению значений.
def find_regex_in_df(df, regex):
    # Список для хранения координат ячеек с совпадениями
    matches = []
    # Проходим по каждой ячейке DataFrame
    for row in range(len(df)): # for row in df.index:
        # for col in df.columns:
        for col in range(len(df.columns)):
            cell_value = df.iat[row, col]
            if pd.isna(cell_value):
                # Если значение ячейки пустое, то переходим к следующей ячейке
                continue
            # Преобразуем значение ячейки в строку
            cell_value = str(cell_value)
            cell_value = delete_symbols_from_string(cell_value, "[().,-]")
            # Если значение ячейки соответствует регулярному выражению
            if re.search(regex, cell_value, re.IGNORECASE):
                # Добавляем координаты ячейки (индекс строки и название столбца) в список совпадений
                matches.append((row, col))
    return matches

# def search_general_bank_info(df) - функция поиска общей информации выписки
# возвращает словарь, состоящий из Искомых заголовков и соответствующих им значений
def search_general_bank_info(df, search_info_correlation) -> Tuple[Dict[str, Any], List[ProcessingError]]:
    """
    Поиск общей информации выписки с обработкой ошибок

    Returns:
    - Dict: Словарь искомых заголовков и соответствующих им значений
    - List[ProcessingError]: Список ошибок обработки
    """
    col_and_value = {}
    errors = []
    try:
        for key, value in search_info_correlation.items():
            # Список координат ячейки искомого заголовка
            header_coordinates = find_regex_in_df(df, key)
            # Если в файле не найден заголовок, то в словарь записываем:
            # "Искомый заголовок": "Заголовок не найден в выписке"
            if not header_coordinates:
                logger.info(f"Для заголовка: {value[0]} не найден соответствующий заголовок в файле!")
                col_and_value[value[0]] = "Заголовок не найден в выписке"
                errors.append(ProcessingError(
                                    code=ResultMessages.WARNING_HEADERS_NOT_CORRECT.status_code,
                                    message=f"Заголовок: {value[0]} не найден в выписке",
                                    severity=ErrorSeverity.WARNING
                                ))
            # Если заголовок найден, берем его первые найденные координаты
            else:
                coordinate = header_coordinates[0]
                # Поиск значения (целиком), соответствующего заголовку, в прилегающих ячейках
                search_df = get_narrow_search_area(df, coordinate)
                col_and_value_cur, search_value_error = search_whole_value_around_header(search_df, value[0], value[1])
                if search_value_error:
                    errors.extend(search_value_error)
                if not col_and_value_cur:
                    # Поиск значения (целиком), соответствующего заголовку, в прилегающих строках и строке заголовка
                    search_df = get_wide_search_area(df, coordinate)
                    col_and_value_cur, search_value_error = search_whole_value_around_header(search_df, value[0], value[1])
                    if search_value_error:
                        errors.extend(search_value_error)
                    if not col_and_value_cur:
                        # Поиск значения (разбитого по ячейкам), соответствующего заголовку, в прилегающих строках и строке заголовка
                        col_and_value_cur, search_value_error = search_value_around_header(search_df, value[0], value[1])
                        if search_value_error:
                            errors.extend(search_value_error)
                        if not col_and_value_cur:
                            # Поиск значения (целиком), соответствующего заголовку, непосредственно в ячейке заголовка
                            # search_df = df.iloc[coordinate[0], coordinate[1]]
                            match = re.search(value[1], df.iloc[coordinate[0], coordinate[1]])
                            if match:
                                col_and_value_cur = {value[0]: match.group()}
                            # Если значение не найдено, то в словарь записываем:
                            # "Искомый заголовок": "Значение не найдено в выписке"
                            if not col_and_value_cur:
                                logger.info(f"Для заголовка: {value[0]} не найдено значение!")
                                # col_and_value[value[0]] = "Значение не найдено в выписке"
                                col_and_value_cur = {value[0]: "Значение не найдено в выписке"}
                                errors.append(ProcessingError(
                                    code=ResultMessages.WARNING_HEADERS_NOT_CORRECT.status_code,
                                    message=f"Значение заголовка: {value[0]} не найдено в выписке",
                                    severity=ErrorSeverity.WARNING
                                ))
                col_and_value.update(col_and_value_cur)
    except Exception as e:
        search_error = ProcessingError(
            code=400,
            message=f"Ошибка поиска общей информации: {str(e)}",
            severity=ErrorSeverity.WARNING,
            details={"exception": str(e)}
        )
        errors.append(search_error)
    return col_and_value, errors

# search_value_around_header функция поиска значения заголовка,
# данные которого находятся в отдельных ячейках
def search_value_around_header(search_area, column_name, search_value_pattern) -> Tuple[Dict[str, Any], List[ProcessingError]]:
    """
    Поиск значения заголовка с обработкой ошибок

    Returns:
    - Dict: Словарь найденных значений
    - List[ProcessingError]: Список ошибок обработки
    """
    column_values = {}
    errors = []
    if column_name == DEBTOR_ACCOUNT_NUMBER:
        seq_len = 20
    else:
        seq_len = 3
    # Проходим по каждой ячейке DataFrame
    try:
        for row in range(len(search_area)):
            # Список для хранения ячеек с совпадениями
            matches = []
            match_num =  []
            for col in range(len(search_area.columns)):
                cell_value = search_area.iat[row, col]
                if pd.isna(cell_value):
                    if match_num:
                        matches.append(match_num)
                    match_num =  []
                    # Если значение ячейки пустое, то переходим к следующей ячейке
                    continue
                # Преобразуем значение ячейки в строку
                cell_value = str(cell_value)
                cell_match_num = re.search(r'^[0-9](?=\.0|$)', cell_value, re.IGNORECASE)
                # Если значение ячейки соответствует регулярному выражению
                if cell_match_num:
                    # Добавляем найденное значение в список совпадений
                    match_num.append(cell_match_num.group())
            matches.append(match_num)
            # Вызываем функцию
            strs_of_length_n = find_lists_of_length_n(matches, seq_len)
            for str_n in strs_of_length_n:
                if re.search(search_value_pattern, str_n, re.IGNORECASE):
                     column_values[column_name] = str_n
    except Exception as e:
        search_value_error = ProcessingError(
            code=400,
            message=f"Ошибка поиска значения: {str(e)}",
            severity=ErrorSeverity.WARNING,
            details={"exception": str(e)}
        )
        errors.append(search_value_error)
    return column_values, errors

# search_whole_value_around_col_name функция поиска значения в ограниченной области
def search_whole_value_around_header(df_search, column_name, search_value_pattern) -> Tuple[Dict[str, Any], List[ProcessingError]]:
    """
    Поиск значения в ограниченной области с обработкой ошибок

    Returns:
    - Dict: Словарь найденных значений
    - List[ProcessingError]: Список ошибок обработки
    """
    column_values = {}
    errors = []
    try:
        # Поиск координат значения по паттерну в переданной области
        cells_coordinates = find_regex_in_df(df_search, search_value_pattern)
        # Если координаты найдены, берем первые и записываем найденное значение в словарь
        if cells_coordinates:
            column_values[column_name] = df_search.iloc[cells_coordinates[0]]
    except Exception as e:
        search_value_error = ProcessingError(
            code=ResultMessages.WARNING_HEADERS_NOT_CORRECT.status_code,
            message=f"Ошибка поиска значения: {str(e)}",
            severity=ErrorSeverity.WARNING,
            details={"exception": str(e)}
        )
        errors.append(search_value_error)
    return column_values, errors

# get_narrow_search_area функция определения области (поиска) вокруг заголовка,
# область - ближайшие ячейки вокруг заголовка
# для последующего поиска соответствующего этому заголовку значения
def get_narrow_search_area(df, header_coordinates):
    row = header_coordinates[0]
    col = header_coordinates[1]
    if len(df) > row == 0 and 0 == col < len(df.columns):
        rows = (row, row + 1)
        cols = (col, col + 1)
    elif len(df) > row == 0 and 0 < col < len(df.columns):
        rows = (row, row + 1)
        cols = (col - 1, col + 1)
    elif len(df) > row == 0 and 0 < col == len(df.columns):
        rows = (row, row + 1)
        cols = (col - 1, col)
    elif len(df) > row > 0 and 0 < col == len(df.columns):
        rows = (row - 1, row + 1)
        cols = (col - 1, col)
    elif len(df) == row > 0 and 0 < col == len(df.columns):
        rows = (row - 1, row)
        cols = (col - 1, col)
    elif len(df) == row > 0 and 0 < col < len(df.columns):
        rows = (row - 1, row)
        cols = (col - 1, col + 1)
    elif len(df) == row > 0 and 0 == col < len(df.columns):
        rows = (row - 1, row)
        cols = (col, col + 1)
    elif len(df) > row > 0 and 0 == col < len(df.columns):
        rows = (row - 1, row + 1)
        cols = (col, col + 1)
    elif len(df) > row > 0 and 0 < col < len(df.columns):
        rows = (row - 1, row + 1)
        cols = (col - 1, col + 1)
    # Однострочный фрейм
    elif len(df) == 1 and 0 == col < len(df.columns):
        rows = (row, row)
        cols = (col, col + 1)
    elif len(df) == 1 and 0 < col < len(df.columns):
        rows = (row, row)
        cols = (col - 1, col + 1)
    elif len(df) == 1 and 0 < col == len(df.columns):
        rows = (row, row)
        cols = (col - 1, col)
    # Одноколоночный фрейм
    elif len(df.columns) == 1 and 0 == row < len(df):
        rows = (row, row + 1)
        cols = (col, col)
    elif len(df.columns) == 1 and 0 < row < len(df):
        rows = (row - 1, row + 1)
        cols = (col, col)
    elif len(df.columns) == 1 and 0 < row == len(df):
        rows = (row - 1, row)
        cols = (col, col)
    # len(df.columns) == 1 and len(df) == 1:
    else:
        rows = (row, row)
        cols = (col, col)
    # Преобразование в срезы
    row_slice = slice(rows[0], rows[1] + 1)
    col_slice = slice(cols[0], cols[1] + 1)
    search_area = df.iloc[row_slice, col_slice]
    return search_area

# get_wide_search_area функция определения области (поиска) вокруг заголовка,
# область - строка выше заголовка, строка заголовка, строка ниже заголовка
# для последующего поиска соответствующего этому заголовку значения
def get_wide_search_area(df, header_coordinates):
    header_row = header_coordinates[0]
    search_area = df.iloc[header_row-1:header_row+2]
    return search_area

# clean_dataframe(df) функция очистки DataFrame от пустых столбцов и строк,
# а также удаление строк, удовлетворяющих указанным значениям
def clean_dataframe(df) -> Tuple[pd.DataFrame, List[ProcessingError]]:
    """
    Очистка DataFrame от пустых столбцов и строк с обработкой ошибок

    Returns:
    - DataFrame: Очищенный DataFrame
    - List[ProcessingError]: Список ошибок обработки
    """
    errors = []
    try:
        # Удаление пустых столбцов
        df =df.dropna(axis=1, how='all')
        # Сбрасываем индекс столбцов, присваивая новые метки
        df.columns = range(df.shape[1])
        # Удаление пустых строк
        df = df.dropna(axis=0, how='all').reset_index(drop=True)
        # Проверяем каждую строку на соответствие последовательности от 1 до общ кол-ва столбцов
        df = clean_sequential_columns_numbers(df)
        # Заменяем \n на пустую строку во всем датафрейме
        pd.set_option('future.no_silent_downcasting', True)
        df = df.replace(r'\n', ' ', regex=True)
    except Exception as e:
        cleaning_error = ProcessingError(
            code=ResultMessages.ERROR_DATAFRAME_CLEANUP_FAILED.status_code,
            message=f"Ошибка очистки DataFrame: {str(e)}",
            severity=ErrorSeverity.WARNING,
            details={"exception": str(e)}
        )
        errors.append(cleaning_error)
    return df, errors

# flatten_headers функция преобразования названий столбцов из многоколоночных в одноколоночные,
# удаление пустых подзаголовков
def flatten_headers(df) -> List[str]:
    # Предположим, что df уже содержит данные, включая две строки заголовков
    # инициализируем новые заголовки пустым списком
    new_headers = []
    # Проверяем первую строку и копируем значения слева направо, если справа NaN.
    for col in range(1, len(df.columns)):  # Начинаем с 1, потому что 0-ой столбец не имеет предшественника
        if pd.isnull(df.iloc[0, col]):
            df.iloc[0, col] = df.iloc[0, col-1]
    # Итерируем по столбцам
    for col in range(len(df.columns)):
        # Для столбцов, где первая строка не является NaN, и вторая строка содержит значение
        if pd.notnull(df.iloc[0, col]) and pd.notnull(df.iloc[1, col]):
            # Объединяем значения из первой и второй строки
            new_header = f"{df.iloc[0, col]} {df.iloc[1, col]}"
        # Для столбцов, где первая строка не является NaN, но вторая строка - NaN
        elif pd.notnull(df.iloc[0, col]):
            # Используем значение из первой строки
            new_header = df.iloc[0, col]
        # Для столбцов, где первая строка - NaN
        else:
            # Используем значение из второй строки (или другую логику, если нужно)
            new_header = df.iloc[1, col]
        new_headers.append(new_header)
    return new_headers

# correct_df_headers функция исправления названий столбцов, полученных из входящих файлов
# в названия переменных, соответствующих столбцам исходящего файла
def correct_df_headers(df, headers_count) -> Tuple[pd.DataFrame, pd.DataFrame, List[ProcessingError]]:
        """
        Корректировка заголовков DataFrame с обработкой ошибок

        Returns:
        - DataFrame: DataFrame с корректированными заголовками
        - List[str]: Оригинальные заголовки
        - List[ProcessingError]: Список ошибок обработки
        """
        errors = []
    # try:
        headers_rows = df.iloc[:headers_count].copy()
        if headers_count == 2:
            original_columns = flatten_headers(headers_rows)
            count_of_row_to_del = 2
        else:
            original_columns = df.iloc[0].tolist()
            count_of_row_to_del = 1

        headers_correlation = {
            SEQ_NUMBER_PATTERN: SEQ_NUMBER,
            OPERATION_DATE_PATTERN: DOCUMENT_OPERATION_DATE,
            DOCUMENT_TYPE_CODE_PATTERN: DOCUMENT_TYPE_CODE,
            DOCUMENT_NUMBER_PATTERN: DOCUMENT_NUMBER,
            DOCUMENT_DATE_PATTERN: DOCUMENT_DATE,
            CORESPONDENT_ACCOUNT_NUMBER_PATTERN: CORESPONDENT_ACCOUNT_NUMBER,
            PAYER_OR_RECIPIENT_BANK_PATTERN: PAYER_OR_RECIPIENT_BANK,
            BANK_BIK_PATTERN: BANK_BIK,
            PAYER_OR_RECIPIENT_NAME_PATTERN: PAYER_OR_RECIPIENT_NAME,
            PAYER_OR_RECIPIENT_INN_PATTERN: PAYER_OR_RECIPIENT_INN,
            PAYER_OR_RECIPIENT_KPP_PATTERN: PAYER_OR_RECIPIENT_KPP,
            ACCOUNT_NUMBER_PATTERN: ACCOUNT_NUMBER,
            DEBIT_AMOUNT_PATTERN: DEBIT_AMOUNT,
            CREDIT_AMOUNT_PATTERN: CREDIT_AMOUNT,
            PAYMENT_PURPOSE_PATTERN: PAYMENT_PURPOSE
        }
        count_of_obligatory_headers = len(headers_correlation)
        obligatory_headers = set(headers_correlation.values())
        # Определение символов для удаления
        chars_to_remove = "[().,-]"

        extra_columns = list(range(len(original_columns)))
        new_columns = original_columns.copy()

        for index, original_col in enumerate(original_columns):
            for key, value in headers_correlation.items():
                cleaned_col = delete_symbols_from_string(original_col, chars_to_remove)
                if re.search(key, cleaned_col, re.IGNORECASE):
                    # Ищем в списке new_columns элемент, равный original_col, и заменяем его по индексу на value
                    new_columns[index] = value
                    del headers_correlation[key]
                    extra_columns.remove(index)
                    break
                else:
                    continue
        # Если все обязательные заголовки найдены в выписке, НО выписка содержит дополнительные столбцы
        # Если НЕ все обязательные заголовки нашлись в выписке (часть найдена, часть нет)
        if ((not headers_correlation and count_of_obligatory_headers < len(original_columns))
                or (headers_correlation and count_of_obligatory_headers > len(headers_correlation))):
            if extra_columns:
            # Удаляем дополнительные столбцы
                try:
                    df = df.drop(columns=extra_columns)
                    # Сбрасываем индекс столбцов, присваивая новые метки
                    df.columns = range(df.shape[1])
                # except Exception as e:
                #     logger.error(f"При удалении дополнительных столбцов возникла ошибка: {e}")
                except Exception as e:
                    header_correction_error = ProcessingError(
                        code=ResultMessages.ERROR_HEADER_CORRECTION_FAILED.status_code,
                        message=ResultMessages.ERROR_HEADER_CORRECTION_FAILED.message,
                        severity=ErrorSeverity.CRITICAL,
                        details={"exception": str(e)}
                    )
                    logger.error(f" При удалении дополнительных столбцов возникла ошибка: {header_correction_error.message}: {str(e)}")
                    errors.append(header_correction_error)
                    return df, pd.DataFrame(), errors
                else:
                    # Фильтрация списка new_columns, чтобы оставить только те элементы, которые есть в obligatory_headers
                    new_columns = [column for column in new_columns if column in obligatory_headers]
            df.columns = new_columns
        # Если все обязательные заголовки найдены в выписке
        elif not headers_correlation and count_of_obligatory_headers == len(new_columns):
            df.columns = new_columns
        # Если НЕ все обязательные заголовки нашлись в выписке (часть найдена, часть нет)
        if headers_correlation:
            # Задаем значения для не найденных заголовков
            col_and_value = {}
            logger.info(f"Заголовки {headers_correlation.values()} не найдены")
            for hd in headers_correlation.values():
                col_and_value[hd] = "Заголовок не найден в выписке"
                errors.append(ProcessingError(
                                    code=ResultMessages.WARNING_HEADERS_NOT_CORRECT.status_code,
                                    message=f"Заголовок {hd} не найден в выписке",
                                    severity=ErrorSeverity.WARNING
                                ))
            for col, val in col_and_value.items():
                    df[col] = val
        # Выбираем строки оригинальных заголовков
        df_with_origin_headers = df.iloc[:count_of_row_to_del].copy()
        # Удаляем первые строки, которые были идентифицированы как заголовки
        df = df.iloc[count_of_row_to_del:].reset_index(drop=True)

        return df, df_with_origin_headers, errors

def clean_headers_in_bank_statement(df, headers_df) -> Tuple[pd.DataFrame, List[ProcessingError]]:
    """
    Удаление заголовков для выписок, представляющих сведения одновременно по нескольким счетам

    Returns:
    - DataFrame: DataFrame с удаленными заголовками
    - List[ProcessingError]: Список ошибок обработки
    """
    errors = []
    try:
        # Объединим df1 и df2 с помощью outer join и добавим индикатор
        merged_df = df.merge(headers_df, how='outer', indicator=True)

        # Фильтруем строки, которые есть только в df1
        filtered_df = merged_df[merged_df['_merge'] == 'left_only'].drop(columns=['_merge'])
    except Exception as e:
        headers_in_bank_statement_error = ProcessingError(
            code=400,
            message=f"Ошибка удаления заголовков: {str(e)}",
            severity=ErrorSeverity.WARNING,
            details={"exception": str(e)}
        )
        errors.append(headers_in_bank_statement_error)
        return df, errors
    return filtered_df, errors

def clean_agregate_rows(df) -> Tuple[pd.DataFrame, List[ProcessingError]]:
    """
    Удаление строк с агрегирующими значениями

    Returns:
    - DataFrame: DataFrame с удаленными строками
    - List[ProcessingError]: Список ошибок обработки
    """
    errors = []
    try:
        # Рассчитываем минимальное количество непропущенных значений для сохранения
        thresh = len(df.columns) // 2 + 1
        # Удаляем строки, заполненные меньше чем на половину
        cleaned_df = df.dropna(thresh=thresh).reset_index(drop=True)
    except Exception as e:
        agregate_rows_error = ProcessingError(
            code=400,
            message=f"Ошибка удаления строк: {str(e)}",
            severity=ErrorSeverity.WARNING,
            details={"exception": str(e)}
        )
        errors.append(agregate_rows_error)
        return df, errors
    return cleaned_df, errors

def clean_sequential_columns_numbers(df):
    # Проверяем каждую строку на соответствие последовательности от 1 до общ кол-ва столбцов
    # Создаем маску для строк, которые нужно удалить
    mask = df.apply(lambda row: np.array_equal(row.values.astype(str), [str(i) for i in range(1, len(row)+1)]), axis=1)
    # Удаление строк, соответствующих маске
    df = df[~mask].reset_index(drop=True)
    return df

# validate_and_log(value, pattern, column_name) функция валидации значения в столбце,
# согласно указанному шаблону
def validate_and_log(value, pattern, column_name):
    if re.match(pattern, str(value)) is None:
        logger.warning(f'Значение "{value}" в столбце {column_name} не прошло валидацию')
        return f'Not valid value: "{value}"'
    return value

# convert_to_float(value) функция преобразования строки в число с плавающей точкой
def convert_to_float(value, column_name):
    # Проверка соответствия значения паттерну "число.число или число" с помощью регулярного выражения
    if re.match(r'^\d+(\.\d+)?$', value):
        return float(value)
    # Проверка соответствия значения паттерну "число-число" с помощью регулярного выражения
    elif re.match(r'^\d+-\d+$', value):
        # Замена дефиса на точку для получения формата числа с плавающей точкой
        converted_value = value.replace('-', '.')
        # Преобразование строки в число с плавающей точкой
        return float(converted_value)
    else:
        logger.warning(f'Значение "{value}" в столбце {column_name} не прошло валидацию')
        return f'Not valid value: "{value}"'

# validate_df_columns(df) функция валидации столбцов DataFrame с приведением к нужному типу и формату
def validate_df_columns(df) -> Tuple[pd.DataFrame, List[ProcessingError]]:
    """
    Валидация столбцов DataFrame с приведением к нужному типу и формату

    Returns:
    - DataFrame: DataFrame с валидированными столбцами
    - List[ProcessingError]: Список ошибок обработки
    """
    errors = []
    for col in df.columns:
        if df[col].dtype == float:
            df[col] = df[col].astype(str)
            # Удаляем десятичную часть, если она равна 0
            df[col] = df[col].apply(lambda x: x.split('.')[0] if x.split('.')[1] == '0' else x)
        else:
            df[col] = df[col].astype(str)
        col_len = 0
        col_pattern = ''
        match col:
            case 'document_type_code':
                col_len = 2
            case 'payer_or_recipient_inn':
                col_pattern = '^[0-9]{10,12}$'
                df[col] = df[col].apply(lambda x: '0' + x if (len(x) == 9 or len(x) == 11) else x)
            case 'payer_or_recipient_kpp':
                col_pattern = '^[0-9]{9}$'
                col_len = 9
            case 'account_number':
                col_pattern = '^[245][0-9]{19}$'
            case 'debit_amount' | 'credit_amount':
                df[col] = df[col].replace('nan', '0')
                df[col] = df[col].replace('-', '0')
                df[col] = df[col].apply(lambda x: convert_to_float(x, col))
            case 'correspondent_account_number':
                col_pattern = '^301[02][145][0-9]{15}$'
            case 'bank_bik':
                col_pattern = '^[012][0-9]{8}$'
                col_len = 9
            case 'document_operation_date' | 'document_date':
                df[col] = pd.to_datetime(df[col], dayfirst=True, infer_datetime_format=True)
            case _:
                print("Other")
        if col_len != 0:
            # Добавляем ведущий ноль, если длина строки меньше положенной
            df[col] = df[col].apply(lambda x: '0' + x if len(x) == col_len-1 else x)
        # Проверка валидации по заданному паттерну
        if col_pattern != '':
            df[col] = df[col].apply(lambda x: validate_and_log(x, col_pattern, col))
    return df, errors

def find_regex_in_df2(df, regex, index):
    presupposed_headers_position = {}
    for i in range(index, -1, -1):
        match_list = []
        for col in range(len(df.columns)):
            cell_value = df.iat[i, col]
            if not pd.isna(cell_value):
                cell_value = str(cell_value)
                cell_value = delete_symbols_from_string(cell_value, "[().,-]")
                cell_match_num = re.search(regex, cell_value, re.IGNORECASE)
                if cell_match_num:
                    match_list.append(1)
                else:
                    match_list.append(0)
        presupposed_headers_position[i] = sum(match_list)
    # Удаление элементов со значением 0
    cleaned_dict = {key: value for key, value in presupposed_headers_position.items() if value != 0}
    return cleaned_dict

# parse_fns_df функция парсинга DataFrame на логические области:
# 1 - Общая информация Выписки
# 2 - Информация об операциях
def parse_fns_df_format(file_df):
    # Пустой список для добавления DataFrame
    found_dfs = []
    # Поиск начала таблицы операций по счету
    search = 'Таблица 1'
    row_idx_of_table1 = file_df.loc[file_df.isin([search]).any(axis=1)].index.tolist()
    if row_idx_of_table1:
        found_dfs.append(file_df.iloc[:row_idx_of_table1[0]])
    # Поиск конца таблицы, содержащей информацию о совершенных операциях по счету
    search = 'Таблица 2'
    row_idx_of_table2 = file_df.loc[file_df.isin([search]).any(axis=1)].index.tolist()
    if row_idx_of_table2:
        found_dfs.append(file_df.iloc[row_idx_of_table1[0]+1:row_idx_of_table2[0]-1])
    # Возврат списка найденных областей
    return found_dfs

# def read_excel функция определения формата выписки
def detect_file_type(df):
    file_format = ''
    found = ()
    search_fns = ['ММВ-7-2/519@','ММВ-7-2/679@']
    for search in search_fns:
        found = df.iloc[:15].apply(lambda row: row.astype(str).str.contains(search), axis=1).any(axis=1)
        if found.any():
            file_format = 'FNS'
            logger.info('found file FNS-format')
            break
    if not found.any():
        logger.info('found file another-format')
    return file_format

def convert_xls_to_xlsx(input_file):
    """
    Конвертирует файл xls в xlsx.

    Args:
        input_file (str): Путь к входному файлу (.xls).

    Returns:успешно
        BytesIO: Данные файла в формате .xlsx в памяти.
    """
    try:
        # Читаем содержимое .xls через pandas
        xls_data = pd.read_excel(input_file, engine='openpyxl')
        new_columns = [
            col if 'UNNAMED' not in col.upper() else ''  # Условие исправления
            for col in xls_data.columns
        ]
        xls_data.columns = new_columns

        # Сохраняем в формате .xlsx
        xlsx_data = BytesIO()
        xls_data.to_excel(xlsx_data, index=False, engine='openpyxl')
        xlsx_data.seek(0)

        logger.info(f"Файл {input_file} успешно конвертирован в формат .xlsx!")
        with open("files/output.xlsx", "wb") as file:
            file.write(xlsx_data.getvalue())

        return xlsx_data
    except Exception as e:
        logger.error(f"Ошибка при конвертации файла {input_file}: {str(e)}")
        raise

def find_lists_of_length_n(nested_lists, n):
    # Используем генератор списков для нахождения всех списков длиной n
    return [''.join(sublist)  for sublist in nested_lists if len(sublist) == n]